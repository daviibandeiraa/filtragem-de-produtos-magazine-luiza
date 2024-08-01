import os
import time
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
def init_driver():
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-infobars')
    chrome_options.add_argument('--start-maximized')
    # chrome_options.add_argument('--headless') 
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def main():
    driver = None
    try:
        driver = init_driver()

        driver.get("https://www.magazineluiza.com.br/")
        search = driver.find_element(By.XPATH, "//input[@type='search']")
        search.send_keys("notebooks\n")
        time.sleep(4)
        notebooks = driver.find_elements(By.XPATH, "//li[@class='sc-SSKRe kzxbRz']")

        # Criar a planilha e as páginas
        book = openpyxl.Workbook()
        book.create_sheet("Piores")
        book.create_sheet("Melhores")

        melhores_page = book['Melhores']
        melhores_page.append(['Produto', 'QTD_Aval', 'URL'])

        piores_page = book['Piores']
        piores_page.append(['Produto', 'QTD_Aval', 'URL'])

        # Processar cada notebook encontrado
        for notebook in notebooks:
            tag_a = notebook.find_element(By.XPATH, "./a")
            redirect = tag_a.get_attribute("href")  # Atribuindo URL correta ao redirect
            info = notebook.find_element(By.XPATH, ".//div[@data-testid='product-card-content']")
            pattern = (
                r"(?P<product>.*)\n"
                r"(?P<rating>[\d.]+) .(?P<assessments>\d+)."
            )
            regex = re.search(pattern, info.text)
            if regex:
                data = regex.groupdict()
                product = data.get("product")
                assessments = data.get("assessments")
                
                # Convert assessments to int for comparison
                try:
                    assessments_int = int(assessments.replace(',', '').replace('.', ''))
                except ValueError:
                    assessments_int = 0
                
                # Adicionar páginas ao arquivo Excel
                if assessments_int < 100:
                    piores_page.append([product, assessments, redirect])
                else:
                    melhores_page.append([product, assessments, redirect])

        # Certificar que a pasta "Output" existe
        output_dir = "Output"
        os.makedirs(output_dir, exist_ok=True)

        # Salvar a planilha na pasta "Output"
        output_path = os.path.join(output_dir, "Notebooks.xlsx")
        book.save(output_path)

    except Exception as e:
        print(f"An error occurred: {e}")
    
    finally:
        if driver:
            driver.quit()

def enviar_email_com_anexo(remetente, destinatario, assunto, corpo, arquivo):
    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = assunto

    # Corpo do e-mail
    msg.attach(MIMEText(corpo, 'plain'))

    # Anexo
    nome_arquivo = arquivo.split("/")[-1]  # Extrai o nome do arquivo
    with open(arquivo, 'rb') as f:
        conteudo = f.read()
    parte = MIMEBase('application', 'octet-stream')
    parte.set_payload(conteudo)
    encoders.encode_base64(parte)
    parte.add_header('Content-Disposition', f"attachment; filename={nome_arquivo}")
    msg.attach(parte)

    # Conectar ao servidor SMTP e enviar o e-mail
    with smtplib.SMTP('smtp-mail.outlook.com', 587) as server:
        server.starttls()
        server.login('cleitin1208@outlook.com', 'Davi0812')
        texto = msg.as_string()
        server.sendmail(remetente, destinatario, texto)

# Exemplo de uso
remetente = 'cleitin1208@outlook.com'
destinatario = 'cleitin1208@outlook.com'
assunto = 'Relatório Notebooks'
corpo = """ 
Olá, aqui está o seu relatório dos notebooks extraídos da Magazine Luiza.

Atenciosamente,
Robô."""
arquivo = './Output/Notebooks.xlsx' 

enviar_email_com_anexo(remetente, destinatario, assunto, corpo, arquivo)

if __name__ == "__main__":
    main()
